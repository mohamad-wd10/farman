"""
Helper utilities for AI Core processing
"""

import re
import pandas as pd
from typing import List, Dict, Any, Tuple


def extract_columns_from_df(df: pd.DataFrame) -> List[str]:
    """
    Extract and clean column names from DataFrame.
    
    Args:
        df: Input DataFrame
        
    Returns:
        List of column names as strings
    """
    columns = []
    for col in df.columns:
        if pd.isna(col):
            columns.append('ستون_خالی')
        else:
            columns.append(str(col).strip())
    return columns


def get_sample_values(df: pd.DataFrame, n: int = 10) -> List[str]:
    """
    Get sample values from the DataFrame for analysis.
    
    Args:
        df: Input DataFrame
        n: Number of samples to extract
        
    Returns:
        List of string representations of sample values
    """
    samples = []
    
    # Sample from first few rows
    for i in range(min(n, len(df))):
        row = df.iloc[i]
        for val in row.values:
            if not pd.isna(val):
                samples.append(str(val))
    
    # Sample from first few columns
    for col in df.columns[:3]:
        col_samples = df[col].dropna().head(n).tolist()
        samples.extend([str(s) for s in col_samples])
    
    return samples[:50]  # Limit to 50 samples


def detect_merge_cells(file_path: str) -> Tuple[bool, List[Dict]]:
    """
    Detect merged cells in an Excel file.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Tuple of (has_merged_cells, list of merge info dicts)
    """
    try:
        # Use openpyxl for detecting merged cells
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path, data_only=True)
        merges_info = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet.merged_cells.ranges:
                for merged_range in sheet.merged_cells.ranges:
                    merges_info.append({
                        'sheet': sheet_name,
                        'range': str(merged_range),
                        'top_left': str(merged_range.min_col) + str(merged_range.min_row),
                        'bottom_right': str(merged_range.max_col) + str(merged_range.max_row)
                    })
        
        return len(merges_info) > 0, merges_info
    
    except Exception as e:
        return False, [{'error': str(e)}]


def is_persian_text(text: str) -> bool:
    """
    Check if text contains Persian characters.
    
    Args:
        text: Input string
        
    Returns:
        True if Persian characters detected
    """
    if not text:
        return False
    
    # Persian Unicode range: \u0600-\u06FF
    persian_pattern = re.compile(r'[\u0600-\u06FF]')
    return bool(persian_pattern.search(text))


def normalize_persian_string(text: str) -> str:
    """
    Normalize Persian string by fixing common character issues.
    
    Args:
        text: Input string
        
    Returns:
        Normalized Persian string
    """
    if not text or not isinstance(text, str):
        return text
    
    # Character replacements
    replacements = {
        'ك': 'ک',  # Arabic Kaf to Persian Kaf
        'ي': 'ی',  # Arabic Ye to Persian Ye
        'ۀ': 'ه',  # He with goal
        'ة': 'ه',  # Ta marbuta
        'آ': 'آ',  # Keep Alef with Madda
        'أ': 'ا',  # Alef with Hamza above
        'إ': 'ا',  # Alef with Hamza below
        'ٰ': '',   # Remove superscript Alef
        '‌‌': '‌',  # Double ZWNJ to single
        '  ': ' ',  # Double space to single
    }
    
    result = text
    for wrong, correct in replacements.items():
        result = result.replace(wrong, correct)
    
    return result.strip()


def estimate_data_quality(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Estimate overall data quality score.
    
    Args:
        df: Input DataFrame
        
    Returns:
        Dict with quality metrics and overall score
    """
    total_cells = df.size
    empty_cells = df.isna().sum().sum()
    
    # Calculate emptiness ratio
    emptiness_ratio = empty_cells / total_cells if total_cells > 0 else 1.0
    
    # Check for duplicate rows
    duplicate_ratio = (len(df) - len(df.drop_duplicates())) / len(df) if len(df) > 0 else 0.0
    
    # Check column name quality
    bad_column_names = 0
    for col in df.columns:
        col_str = str(col)
        if pd.isna(col) or col_str.strip() == '' or col_str.startswith('Unnamed'):
            bad_column_names += 1
    
    column_quality = 1.0 - (bad_column_names / len(df.columns)) if len(df.columns) > 0 else 0.0
    
    # Calculate overall score (0-100)
    quality_score = (
        (1.0 - emptiness_ratio) * 40 +
        (1.0 - duplicate_ratio) * 30 +
        column_quality * 30
    )
    
    return {
        'total_cells': total_cells,
        'empty_cells': int(empty_cells),
        'emptiness_ratio': round(emptiness_ratio, 3),
        'duplicate_ratio': round(duplicate_ratio, 3),
        'column_quality': round(column_quality, 3),
        'quality_score': round(quality_score, 1),
        'rating': get_quality_rating(quality_score)
    }


def get_quality_rating(score: float) -> str:
    """Convert quality score to human-readable rating"""
    if score >= 90:
        return 'عالی'
    elif score >= 75:
        return 'خوب'
    elif score >= 60:
        return 'متوسط'
    elif score >= 40:
        return 'ضعیف'
    else:
        return 'نیاز به بازنگری'


def parse_persian_date(date_str: str) -> Tuple[int, int, int]:
    """
    Parse Persian date string to (year, month, day) tuple.
    
    Args:
        date_str: Date string in various Persian formats
        
    Returns:
        Tuple of (year, month, day) or None if parsing fails
    """
    if not date_str:
        return None
    
    # Common Persian date patterns
    patterns = [
        r'(\d{4})/(\d{1,2})/(\d{1,2})',  # 1402/01/15
        r'(\d{4})-(\d{1,2})-(\d{1,2})',  # 1402-01-15
        r'(\d{2})/(\d{2})/(\d{4})',      # 01/15/1402
    ]
    
    for pattern in patterns:
        match = re.search(pattern, date_str)
        if match:
            groups = match.groups()
            if len(groups[0]) == 4:  # YYYY/MM/DD
                return int(groups[0]), int(groups[1]), int(groups[2])
            elif len(groups[2]) == 4:  # DD/MM/YYYY or MM/DD/YYYY
                # Assume DD/MM/YYYY for Persian dates
                return int(groups[2]), int(groups[1]), int(groups[0])
    
    return None
